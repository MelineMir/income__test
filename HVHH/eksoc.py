import os
from pdfreader import SimplePDFViewer
from openpyxl import Workbook

pdf_folder = 'path_to_folder_containing_pdfs'
excel_file = 'EK.xlsx'

wb = Workbook()

for filename in os.listdir(pdf_folder):
    if filename.endswith('.pdf'):
        pdf_path = os.path.join(pdf_folder, filename)
        ws = wb.create_sheet(title=filename)

        with open(pdf_path, 'rb') as pdf_file:
            pdf_viewer = SimplePDFViewer(pdf_file)

            text = ''
            while pdf_viewer.next():
                text += pdf_viewer.current_text

        ws['A1'] = text

default_sheet_name = 'Sheet'
if default_sheet_name in wb.sheetnames:
    default_sheet = wb[default_sheet_name]
    wb.remove(default_sheet)

wb.save(excel_file)
